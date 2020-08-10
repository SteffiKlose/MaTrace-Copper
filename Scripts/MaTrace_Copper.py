# -*- coding: utf-8 -*-
"""
Created on Fri Mar  9 10:49:33 2018

@author: sklose
"""

import logging
import os
import numpy as np
import time
import datetime
import scipy.io
import scipy.stats   
import matplotlib.pyplot as plt    
import pandas as pd
import shutil 
import uuid
import xlrd


from scipy.optimize import curve_fit
from scipy.integrate import simps
from numpy import trapz
from openpyxl import load_workbook

#%% Define loggig and main Paths
def function_logger(file_level, Name_Scenario, Path_Result, console_level): # Set up logger
    # remove all handlers from logger
    logger = logging.getLogger()
    logger.handlers = [] # required if you don't want to exit the shell
    logger.setLevel(logging.DEBUG) #By default, logs all messages

    if console_level != None:
        console_log = logging.StreamHandler() #StreamHandler logs to console
        console_log.setLevel(console_level)
        console_log_format = logging.Formatter('%(message)s') # ('%(asctime)s - %(message)s')
        console_log.setFormatter(console_log_format)
        logger.addHandler(console_log)

    file_log = logging.FileHandler(str(Path_Result) + '\\' + Name_Scenario + '.html', mode='w', encoding=None, delay=False)
    file_log.setLevel(file_level)
    #file_log_format = logging.Formatter('%(asctime)s - %(lineno)d - %(levelname)-8s - %(message)s<br>')
    file_log_format = logging.Formatter('%(message)s<br>')
    file_log.setFormatter(file_log_format)
    logger.addHandler(file_log)
    return logger, console_log, file_log
      
def ensure_dir(f): # Checks whether a given directory f exists, and creates it if not
    d = os.path.dirname(f)
    if not os.path.exists(d):
        os.makedirs(d) 
        
def greyfade(color_triple,greyfactor): # returns a color triple (RGB) that is between the original color triple and and equivalent grey value, linear scaling: greyfactor = 0: original, greyfactor = 1: grey
    return (1 - greyfactor) * color_triple + greyfactor * color_triple.sum() / 3


"""
Set main path
"""    

from pathlib import Path

# `cwd`: current directory
cwd = Path.cwd()
    
#Path_List=[]
#Input_Paths = open('Path_names.txt','r')
#for line in Input_Paths.read().split('\n'):
#    Path_List.append(line)
#Input_Paths.close()    
    
    
Project_MainPath = cwd.parent
Name_User        = '' # Enter your name here
Input_Data        = 'MaTraceCopper_Indata.xlsx'
Path_Data   = Path.joinpath(Project_MainPath ,'Data\\')
Path_Results = Path.joinpath(Project_MainPath ,'Results\\')
Path_Script = Path.joinpath(Project_MainPath ,'Scripts\\')


#%% Read Configuarion file    

Project_DataFileName = Input_Data
Project_DataFilePath = Path.joinpath(Path_Data , Project_DataFileName)
Project_DataFile_WB  = xlrd.open_workbook(Project_DataFilePath)
Project_Configsheet  = Project_DataFile_WB.sheet_by_name('Scenario_Overview')
                
Name_Scenario      = Project_Configsheet.cell_value(5,2)
Number_Scenario      = Project_Configsheet.cell_value(3,2)
StartTime          = datetime.datetime.now()
TimeString         = str(StartTime.year) + '_' + str(StartTime.month) + '_' + str(StartTime.day) + '__' + str(StartTime.hour) + '_' + str(StartTime.minute) + '_' + str(StartTime.second)
Path_Result        = Path.joinpath(Path_Results , Name_Scenario + '_' + TimeString + '\\')

# Read control and selection parameters into dictionary
ScriptConfig = {'Scenario_Description': Project_Configsheet.cell_value(6,2)}
ScriptConfig['Scenario Name'] = Name_Scenario
for m in range(9,35): # add all defined control parameters to dictionary
    try:
        ScriptConfig[Project_Configsheet.cell_value(m,1)] = np.int(Project_Configsheet.cell_value(m,2))
    except:
        ScriptConfig[Project_Configsheet.cell_value(m,1)] = str(Project_Configsheet.cell_value(m,2))
        
ScriptConfig['Current_UUID'] = str(uuid.uuid4())

# Create scenario folder
ensure_dir(Path_Result)
#Copy script and Config file into that folder
os.mkdir(Path_Result)
shutil.copyfile(Project_DataFilePath, Path.joinpath(Path_Result , Project_DataFileName))
shutil.copyfile(Path.joinpath(Project_MainPath , 'Scripts\\MaTrace_Copper.py'), Path.joinpath(Path_Result ,'MaTrace_Copper.py'))
# Initialize logger    
[Mylog,console_log,file_log] = function_logger(logging.DEBUG, Name_Scenario + '_' + TimeString, Path_Result, logging.DEBUG) 

# log header and general information
Mylog.info('<html>\n<head>\n</head>\n<body bgcolor="#ffffff">\n<br>')
Mylog.info('<font "size=+5"><center><b>Script ' + 'MaTrace_Copper' + '.py</b></center></font>')
Mylog.info('<font "size=+5"><center><b>Version: 2020-02-05 or later.</b></center></font>')
Mylog.info('<font "size=+4"> <b>Current User: ' + Name_User + '.</b></font><br>')
Mylog.info('<font "size=+4"> <b>Current Path: ' + str(Project_MainPath) + '.</b></font><br>')
Mylog.info('<font "size=+4"> <b>Current Scenario: ' + Name_Scenario + '.</b></font><br>')
Mylog.info(ScriptConfig['Scenario_Description'])
Mylog.info('Unique ID of scenario run: <b>' + ScriptConfig['Current_UUID'] + '</b>')

Time_Start = time.time()
Mylog.info('<font "size=+4"> <b>Start of simulation: ' + time.asctime() + '.</b></font><br>')


#%% Read model parameters        

Mylog.info('<p><b>Reading model definitions.</b></p>')
Project_DefSheet  = Project_DataFile_WB.sheet_by_name('Definitions')

Par_NoOfProducts  = int(Project_DefSheet.cell_value(2,5))
Mylog.info('<p>Number of products: ' + str(Par_NoOfProducts) + '.</p>')

Par_NoOfProductGroups  = int(Project_DefSheet.cell_value(2,3))
Mylog.info('<p>Number of product groups: ' + str(Par_NoOfProductGroups) + '.</p>')

Par_NoOfYears     = int(Project_DefSheet.cell_value(2,2)) #[2015;2100]
Mylog.info('<p>Number of years: ' + str(Par_NoOfYears) + '.</p>')

Par_NoOfScraps    = int(Project_DefSheet.cell_value(2,7))
Mylog.info('<p>Number of scrap types: ' + str(Par_NoOfScraps) + '.</p>')

Par_NoOfRecyclingRoutes   = int(Project_DefSheet.cell_value(2,6)) 
Mylog.info('<p>Number of Recycling Routes: ' + str(Par_NoOfRecyclingRoutes) + '.</p>')

Par_NoOfSecMetals = int(Project_DefSheet.cell_value(2,8)) 
Mylog.info('<p>Number of refinement processes: ' + str(Par_NoOfSecMetals) + '.</p>')

Par_NoOfRegions   = int(Project_DefSheet.cell_value(2,9)) 
if Par_NoOfRegions==1:
    Mylog.info('Number of Regions: global ' + '.</p>')
else: Mylog.info('<p>Number of Regions: ' + str(Par_NoOfRegions) + '.</p>')


Def_ProductNames      = []
for m in range (0,Par_NoOfProducts):
    Def_ProductNames.append(Project_DefSheet.cell_value(m+4,5))    

Par_Time = [] # Time vector of the model, unit: year, first element: 2015
for m in range(0,Par_NoOfYears):
    Par_Time.append(np.int(Project_DefSheet.cell_value(m+4,2)))
    
if Par_NoOfRegions==1:
    Project_Datasheet  = Project_DataFile_WB.sheet_by_name('Parameters_global')
    
else:
    Project_Datasheet  = Project_DataFile_WB.sheet_by_name('Parameters_regions')
   
Mylog.info('Read and format parameters for manufacture and use phase.<br>')

#%% Read Efficiency Parameters

if ScriptConfig['Modus'] == 'Trace 2015 mined copper': 
    Global_Copper_Production2015 = Project_Datasheet.cell_value(5,2)  
   
Copper_use_in_regions_share=np.zeros(Par_NoOfRegions)
for r in range(0,Par_NoOfRegions):
    Copper_use_in_regions_share[r]=Project_Datasheet.cell_value(5+(Par_NoOfProducts+1)*r,3)
    
Copper_use_in_regions=np.zeros(Par_NoOfRegions)
for r in range(0,Par_NoOfRegions):
    Copper_use_in_regions[r]=Project_Datasheet.cell_value(5+(Par_NoOfProducts+1)*r,3)*Global_Copper_Production2015
    
Par_D_AllocationCopperToProducts = np.zeros((Par_NoOfRegions,Par_NoOfProducts)) # This is the D-matrix
for r in range(0,Par_NoOfRegions):
    for p in range(0,Par_NoOfProducts):
           Par_D_AllocationCopperToProducts[r,p]  = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*r,7)

    
Par_Input_F_0_8 = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))  #Fabrication yield in regions
Par_Input_F_0_8[0,:,:] = np.einsum('ij,i->ij',Par_D_AllocationCopperToProducts[:,:],Copper_use_in_regions[:]) 

for m in range (1,Par_NoOfYears):
    for r in range(0,Par_NoOfRegions):
        for p in range(0,Par_NoOfProducts):
            Par_Input_F_0_8[m,:,:] = 0   
            
            
Par_Lambda_Fabrication_Efficiency = np.zeros((Par_NoOfProducts,Par_NoOfRegions))    # fabrication yield for products
for p in range(0,Par_NoOfProducts):
   for r in range(0,Par_NoOfRegions):
       Par_Lambda_Fabrication_Efficiency[p,r] = Project_Datasheet.cell_value(5+p+(Par_NoOfProducts+1)*r,12)

if ScriptConfig['Sensitivity'] == 'Sens fabrication eff long':
    Mylog.info('Sens fabrication eff  short.<br>')
    for p in range(0,Par_NoOfProducts):
        for r in range(0,Par_NoOfRegions): 
                Par_Lambda_Fabrication_Efficiency[p,r] = Par_Lambda_Fabrication_Efficiency[p,r] * 1.1
                if Par_Lambda_Fabrication_Efficiency[p,r] > 1:
                    Par_Lambda_Fabrication_Efficiency[p,r] = 1

if ScriptConfig['Sensitivity'] == 'Sens fabrication eff short':
    Mylog.info('Sens fabrication eff  short.<br>')
    for p in range(0,Par_NoOfProducts):
        for r in range(0,Par_NoOfRegions): 
                Par_Lambda_Fabrication_Efficiency[p,r] = Par_Lambda_Fabrication_Efficiency[p,r] * 0.9


Par_Tau = np.zeros((Par_NoOfRegions,Par_NoOfProducts)) # Mean lifetime array, unit: year
for r in range(0,Par_NoOfRegions):
    for p in range(0,Par_NoOfProducts):
        Par_Tau[r,p] = Project_Datasheet.cell_value(5+p+(Par_NoOfProducts+1)*r,13)

#increase lifetime if lifetime flag is set:
if ScriptConfig['Lifetime extension'] == 'all products':
    Par_Tau = 1.2 * Par_Tau
    Mylog.info('Lifetime extention 20%.<br>')
#read standart deviaion of lifetime
Par_StandDev_Tau=float(Project_Datasheet.cell_value(2,13))

if ScriptConfig['Lifetime extension'] == 'all products in IC':
    for p in range(0,Par_NoOfProducts):
        Par_Tau[0,p] = 1.2 * Par_Tau[0,p]    
        Par_Tau[4,p] = 1.2 * Par_Tau[4,p] 
        Par_Tau[5,p] = 1.2 * Par_Tau[5,p] 
        
if ScriptConfig['Lifetime extension'] == 'C&E':
    for r in range(0,Par_NoOfRegions):
        for p in range(12,16):
            Par_Tau[r,p] = 1.2 * Par_Tau[r,p]  
        
if ScriptConfig['Lifetime extension'] == 'C&E in IC':
   for p in range(12,16):
           Par_Tau[0,p] = 1.2 * Par_Tau[0,p]    
           Par_Tau[4,p] = 1.2 * Par_Tau[4,p] 
           Par_Tau[5,p] = 1.2 * Par_Tau[5,p] 
           
if ScriptConfig['Lifetime extension'] == 'Decreased C&E':
  for r in range(0,Par_NoOfRegions):
    for p in range(12,16):
           Par_Tau[r,p] = 0.8 * Par_Tau[r,p]    
          
if ScriptConfig['Lifetime extension'] == 'Moderate':
  for r in range(0,Par_NoOfRegions):
    for p in range(12,16):
           Par_Tau[r,p] = 2 * Par_Tau[r,p]    
           
if ScriptConfig['Lifetime extension'] == 'Ambitious':
  for r in range(0,Par_NoOfRegions):
    for p in range(12,16):
           Par_Tau[r,p] = 3.5 * Par_Tau[r,p]    

if ScriptConfig['Sensitivity'] == 'Sens Lifetime long':
    Par_Tau = 1.1 * Par_Tau
    Mylog.info('sens lifetime long.<br>')

if ScriptConfig['Sensitivity'] == 'Sens Lifetime short':
    Par_Tau = 0.9 * Par_Tau
    Mylog.info('Sens lifetime short.<br>')

Par_Sigma = Par_StandDev_Tau * Par_Tau # Standard deviation of lifetime array, region by product, unit: year       

Mylog.info('Define lifetime distribution of product cohorts in MaTrace_pdf Matrix.<br>')
MaTrace_pdf = np.zeros((Par_NoOfYears,Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts)) 
AgeMatrix   = np.zeros((Par_NoOfYears,Par_NoOfYears))
for c in range(0, Par_NoOfYears):  # cohort index
    for y in range(c + 1, Par_NoOfYears):
        AgeMatrix[y,c] = y-c
for R in range(0,Par_NoOfRegions):
    for P in range(0,Par_NoOfProducts):
        MaTrace_pdf[:,:,R,P] = scipy.stats.truncnorm((0-Par_Tau[R,P])/Par_Sigma[R,P], np.inf, loc=Par_Tau[R,P], scale=Par_Sigma[R,P]).pdf(AgeMatrix)  # Call scipy's Norm function with Mean, StdDev, and Age
# No discard in historic years and year 0:
for m in range(0,Par_NoOfYears):
    MaTrace_pdf[0:m+1,m,:,:] = 0

Mylog.info('Read parameters for obsolete products and waste management industries.<br>')
Par_Omega_ObsoleteStocks = np.zeros((Par_NoOfRegions,Par_NoOfProducts)) # Par: Ω
for r in range(0,Par_NoOfRegions):
    for p in range(0,Par_NoOfProducts):
        Par_Omega_ObsoleteStocks[r,p] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*r,14)
   
if ScriptConfig['Sensitivity'] == 'Sens omega high':
    Mylog.info('Sens omega high.<br>')
    for r in range(0,Par_NoOfRegions):
        for p in range(0,Par_NoOfProducts):
             Par_Omega_ObsoleteStocks[r,p] = Par_Omega_ObsoleteStocks[r,p] * 1.1
            
if ScriptConfig['Sensitivity'] == 'Sens omega low':
    Mylog.info('Sens omega low.<br>')
    for r in range(0,Par_NoOfRegions):
      for p in range(0,Par_NoOfProducts):
           Par_Omega_ObsoleteStocks[r,p] = Par_Omega_ObsoleteStocks[r,p] * 0.9

Par_Sigma_Losses = np.zeros((Par_NoOfRegions,Par_NoOfProducts)) # Par: Ω
for r in range(0,Par_NoOfRegions):
      for p in range(0,Par_NoOfProducts):
            Par_Sigma_Losses[r,p] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*r,15)

if ScriptConfig['Sensitivity'] == 'Sens sigma high':
    Mylog.info('Sens sigma high.<br>')
    for r in range(0,Par_NoOfRegions):
        for p in range(0,Par_NoOfProducts):
              Par_Sigma_Losses[r,p] = Par_Sigma_Losses[r,p] * 1.1


if ScriptConfig['Sensitivity'] == 'sens sigma low':
    Mylog.info('sens sigma low.<br>')
    for m in range(0,Par_NoOfYears):
        for r in range(0,Par_NoOfRegions):
            for p in range(0,Par_NoOfProducts):
                Par_Sigma_Losses[m,r,p] = Par_Sigma_Losses[m,r,p] * 0.9

Par_Gamma_EoL_Collection_Rate_Copper = np.zeros((Par_NoOfRegions,Par_NoOfProducts))      
for r in range(0,Par_NoOfRegions):
    for p in range(0,Par_NoOfProducts):
          Par_Gamma_EoL_Collection_Rate_Copper[r,p] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*r,16)


                    
#%% Read Process Transformation Parameters

# Allocation of EoL Products to scrap groups in different regions 
Par_A_EolToScrap_Copper = np.zeros((Par_NoOfRegions,Par_NoOfProducts,Par_NoOfScraps))        
for p in range(0,Par_NoOfProducts):
   for r in range(0,Par_NoOfRegions):
      for s in range(0,Par_NoOfScraps):
          Par_A_EolToScrap_Copper[r,p,s] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*r,18+s)
                
if ScriptConfig['WEEE consumer sorting'] == 'Improved':
   for p in range(0,Par_NoOfProducts):
      for r in range(0,Par_NoOfRegions):
          for s in range(0,Par_NoOfScraps):
              Par_A_EolToScrap_Copper[r,p,s] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*Par_NoOfRegions,18+s)
                
if ScriptConfig['WEEE consumer sorting'] == 'Improved in IC':
   for p in range(0,Par_NoOfProducts):
       for s in range(0,Par_NoOfScraps):
           Par_A_EolToScrap_Copper[0,p,s] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*Par_NoOfRegions,18+s)
           Par_A_EolToScrap_Copper[4,p,s] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*Par_NoOfRegions,18+s)
           Par_A_EolToScrap_Copper[5,p,s] = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*Par_NoOfRegions,18+s)
                  
Mylog.info('Read parameters for scrap treatment and refining.<br>')  

Par_Phi_Scrap_Sorting_Efficiency = np.zeros((Par_NoOfRegions,Par_NoOfScraps))  
for r in range(0,Par_NoOfRegions):
    for s in range(0,Par_NoOfScraps):
        Par_Phi_Scrap_Sorting_Efficiency[r,s] = Project_Datasheet.cell_value(s+5+(Par_NoOfProducts+1)*r,27)


                    
                    
if ScriptConfig['Increased sorting rate'] == 'Moderate':
    Mylog.info('Sens scrap sorting efficiency high.<br>')
    for r in range(0,Par_NoOfRegions):
        for s in range(0,Par_NoOfScraps):
                Par_Phi_Scrap_Sorting_Efficiency[r,s] = Par_Phi_Scrap_Sorting_Efficiency[r,s] * 1.1
                if Par_Phi_Scrap_Sorting_Efficiency[r,s] > 1:
                    Par_Phi_Scrap_Sorting_Efficiency[r,s]=1             
                    
                    
if ScriptConfig['Increased sorting rate'] == 'Ambitious':
    Mylog.info('Sens scrap sorting efficiency high.<br>')
    for r in range(0,Par_NoOfRegions):
        for s in range(0,Par_NoOfScraps):
                Par_Phi_Scrap_Sorting_Efficiency[r,s] = Par_Phi_Scrap_Sorting_Efficiency[r,s] * 1.3
                if Par_Phi_Scrap_Sorting_Efficiency[r,s] > 1:
                    Par_Phi_Scrap_Sorting_Efficiency[r,s]=1 
               
if ScriptConfig['Sensitivity'] == 'Sens scrap sorting efficiency low':
    Mylog.info('Sens scrap sorting efficiency low.<br>')
    for r in range(0,Par_NoOfRegions):
        for s in range(0,Par_NoOfScraps):
                Par_Phi_Scrap_Sorting_Efficiency[r,s] = Par_Phi_Scrap_Sorting_Efficiency[r,s] *  0.9


if ScriptConfig['Sensitivity'] == 'Sens scrap sorting efficiency high':
    Mylog.info('Sens scrap sorting efficiency high.<br>')
    for r in range(0,Par_NoOfRegions):
        for s in range(0,Par_NoOfScraps):
                Par_Phi_Scrap_Sorting_Efficiency[r,s] = Par_Phi_Scrap_Sorting_Efficiency[r,s] * 1.1
                if Par_Phi_Scrap_Sorting_Efficiency[r,s] > 1:
                    Par_Phi_Scrap_Sorting_Efficiency[r,s]=1

Par_Theta_Copper_recovery_from_scrap_in_recyclingroute = np.zeros((Par_NoOfRegions,Par_NoOfScraps,Par_NoOfRecyclingRoutes))
for r in range(0,Par_NoOfRegions):
   for s in range(0,Par_NoOfScraps):   
       for t in range (0,Par_NoOfRecyclingRoutes):
          Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] = Project_Datasheet.cell_value(s+5+(Par_NoOfProducts+1)*r,34+t)        
 

if ScriptConfig['Sensitivity'] == 'Sens recovery high':
    Mylog.info('Sens copper recovery high.<br>')
    for r in range(0,Par_NoOfRegions):
        for s in range(0,Par_NoOfScraps):   
            for t in range (0,Par_NoOfRecyclingRoutes):
                Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] = Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] * 1.1
                if Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] > 1:
                    Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] = 1
                    
                    
if ScriptConfig['Sensitivity'] == 'Sens recovery low':
    Mylog.info('Sens recovery low.<br>')
    for r in range(0,Par_NoOfRegions):
        for s in range(0,Par_NoOfScraps):   
            for t in range (0,Par_NoOfRecyclingRoutes):
                Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] = Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[r,s,t] * 0.9

        
Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf= np.zeros((Par_NoOfRegions))
for r in range(0,Par_NoOfRegions):
          Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r] = Project_Datasheet.cell_value(5+(Par_NoOfProducts+1)*r,54)        


if ScriptConfig['IT efficiency'] == 'Moderate':
    Mylog.info('IT efficiency Moderate.<br>')
    for r in range(0,Par_NoOfRegions):
            Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r]=0.8
          
                    
if ScriptConfig['IT efficiency'] == 'Ambitious':
    Mylog.info('IT efficiency Ambitious.<br>')
    for r in range(0,Par_NoOfRegions):
            Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r]=0.95
    
Par_B_ScrapToRemeltingRoute = np.zeros((Par_NoOfRegions,Par_NoOfScraps,Par_NoOfRecyclingRoutes)) # Par: B2- scrap to recycling route 
for r in range(0,Par_NoOfRegions):
    for s in range(0,Par_NoOfScraps):
        for t in range(0,Par_NoOfRecyclingRoutes):
            Par_B_ScrapToRemeltingRoute[r,s,t] = Project_Datasheet.cell_value(s+5+(Par_NoOfProducts+1)*r,t+29) 
                    
Par_inf_collection_rate = np.zeros((Par_NoOfRegions,Par_NoOfProducts))  ##defines how much of EoL Products from Consumer and Electronic sector are collected informally
for r in range(0,Par_NoOfRegions):
   for p in range(0,Par_NoOfProducts):   
       Par_inf_collection_rate[r,p]  = Project_Datasheet.cell_value(p+5+(Par_NoOfProducts+1)*r,48)
       
Par_Phi_inf_Scrap_Sorting_Efficiency = np.zeros((Par_NoOfRegions))  ##defines the manual dissasembly efficiecy of the informal recycling sector
for r in range(0,Par_NoOfRegions):  
       Par_Phi_inf_Scrap_Sorting_Efficiency[r]  = Project_Datasheet.cell_value(5+(Par_NoOfProducts+1)*r,52)
      
if ScriptConfig['Sensitivity'] == 'Sens informal copper recovery high':
    Mylog.info('Sens informal copper recovery high.<br>')
    for r in range(0,Par_NoOfRegions): 
                Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r]=Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r] * 1.3

if ScriptConfig['Sensitivity'] == 'Sens informal copper recovery low':
    Mylog.info('Sens incineration rate low.<br>')
    for r in range(0,Par_NoOfRegions): 
                Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r]=Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[r] * 0.7


if ScriptConfig['Sensitivity'] == 'Sens informal collection rate high':
    Mylog.info('Sens informal collection rate high.<br>')
    for r in range(0,Par_NoOfRegions): 
        for p in range(0, Par_NoOfProducts):
                Par_inf_collection_rate[r,p]=Par_inf_collection_rate[r,p] * 1.3
                if Par_Gamma_EoL_Collection_Rate_Copper[r,p] + Par_inf_collection_rate[r,p] > 1:
                    Par_inf_collection_rate[r,p]=1-Par_Gamma_EoL_Collection_Rate_Copper[r,p]
                    
                if Par_Gamma_EoL_Collection_Rate_Copper[r,p] + Par_inf_collection_rate[r,p] > 1:
                    Par_inf_collection_rate[r,p]=1-Par_Gamma_EoL_Collection_Rate_Copper[r,p]

if ScriptConfig['Sensitivity'] == 'Sens informal collection rate low':
    Mylog.info('Sens informal collection rate low.<br>')
    for r in range(0,Par_NoOfRegions): 
                Par_inf_collection_rate[r,:]=Par_inf_collection_rate[r,:] *  0.7
                
if ScriptConfig['Sensitivity'] == 'Sens Informal scrap sorting efficiency  low':
    Mylog.info('Informal scrap sorting efficiency low.<br>')
    for r in range(0,Par_NoOfRegions): 
                Par_Phi_inf_Scrap_Sorting_Efficiency[r]=Par_Phi_inf_Scrap_Sorting_Efficiency[r] * 0.7


if ScriptConfig['Sensitivity'] == 'Sens Informal scrap sorting efficiency  high':
    Mylog.info('Informal scrap sorting efficiency high.<br>')
    for r in range(0,Par_NoOfRegions): 
                Par_Phi_inf_Scrap_Sorting_Efficiency[r]=Par_Phi_inf_Scrap_Sorting_Efficiency[r] * 1.3
                if Par_Phi_inf_Scrap_Sorting_Efficiency[r] > 1:
                    Par_Phi_inf_Scrap_Sorting_Efficiency[r]=1
                
Par_C_RemeltingToSecondaryMetal = np.zeros((Par_NoOfRegions,Par_NoOfRecyclingRoutes,Par_NoOfSecMetals)) # Par: Theta2, Allocation of Remelted Scrap to Secondary Metals
for r in range(0,Par_NoOfRegions):
        for t in range(0,Par_NoOfRecyclingRoutes):
            for c in range(0,Par_NoOfSecMetals):
                Par_C_RemeltingToSecondaryMetal[r,t,c]  = Project_Datasheet.cell_value(t+5+(Par_NoOfProducts+1)*r,40+c)
 
if ScriptConfig['Sensitivity'] == 'Sens Collection rate high':
    Mylog.info('Sens collection rate high.<br>')
    for r in range(0,Par_NoOfRegions):
        for p in range(0,12):
                Par_Gamma_EoL_Collection_Rate_Copper[r,p] = Par_Gamma_EoL_Collection_Rate_Copper[r,p] * 1.1
        for p in range(12,16):
                Par_Gamma_EoL_Collection_Rate_Copper[r,p] = Par_Gamma_EoL_Collection_Rate_Copper[r,p] * 1.2
              
        if Par_Gamma_EoL_Collection_Rate_Copper[r,p] > 1:
                Par_Gamma_EoL_Collection_Rate_Copper[r,p]=1
                
if ScriptConfig['Sensitivity'] == 'Sens Collection rate low':
    Mylog.info('Sens collection rate low.<br>')
    for r in range(0,Par_NoOfRegions):
        for p in range(0,12):
                Par_Gamma_EoL_Collection_Rate_Copper[r,p] = Par_Gamma_EoL_Collection_Rate_Copper[r,p] * 0.9
        for p in range(12,16):
                Par_Gamma_EoL_Collection_Rate_Copper[r,p] = Par_Gamma_EoL_Collection_Rate_Copper[r,p] * 0.8
                             
 # Improved WEEE  recovery efficiency (Γ) if WEEE Collection rate is set to improve in all regions:               
if ScriptConfig['WEEE collection rate'] == 'Ambitious':       
    Mylog.info('WEEE collection rate Ambitious.<br>')
    for p in range(12,16):
          for r in range(0,Par_NoOfRegions):
                 Par_Gamma_EoL_Collection_Rate_Copper[r,p] =0.85
                 
                 
                 if Par_Gamma_EoL_Collection_Rate_Copper[r,p] > 1:
                    Par_Gamma_EoL_Collection_Rate_Copper[r,p]=1
                 
                 if Par_Gamma_EoL_Collection_Rate_Copper[r,p] + Par_inf_collection_rate[r,p] > 1:
                    Par_inf_collection_rate[r,p]=1-Par_Gamma_EoL_Collection_Rate_Copper[r,p]
       

 # Improved WEEE  recovery efficiency (Γ) if WEEE Collection rate is set to improved in IC:               
if ScriptConfig['WEEE collection rate'] == 'Moderate':       
    Mylog.info('WEEE collection rate Moderate.<br>')
    for r in range(0,Par_NoOfRegions):
            for p in range(12,16):
                Par_Gamma_EoL_Collection_Rate_Copper[r,p] = Par_Gamma_EoL_Collection_Rate_Copper[r,p] * 1.1
                
                if Par_Gamma_EoL_Collection_Rate_Copper[r,p] > 1:
                    Par_Gamma_EoL_Collection_Rate_Copper[r,p]=1
              
                if Par_Gamma_EoL_Collection_Rate_Copper[r,p] + Par_inf_collection_rate[r,p] > 1:
                    Par_inf_collection_rate[r,p]=1-Par_Gamma_EoL_Collection_Rate_Copper[r,p]



#%% Read reuse and trade data
###Reuse von C&E (consumer goods: 12) und Vehicles (non-electrical vehicles: 10)  
          
Mylog.info('Read parameters for ELV trade and WEEE trade.<br>')  

Project_Tradesheet  = Project_DataFile_WB.sheet_by_name('ReuseandTrade')

Par_Chi_Reuse = np.zeros((Par_NoOfProducts,Par_NoOfRegions,Par_NoOfRegions)) 
for p in range(12,16):
   for r in range(0,Par_NoOfRegions):
       for R in range(0,Par_NoOfRegions):
           Par_Chi_Reuse[p,r,R] = Project_Tradesheet.cell_value(r+2,2+R)

Par_Chi_Reuse_inf = np.zeros((Par_NoOfProducts,Par_NoOfRegions,Par_NoOfRegions)) 
for p in range(12,16):
   for r in range(0,Par_NoOfRegions):
       for R in range(0,Par_NoOfRegions):
           Par_Chi_Reuse_inf[p,r,R] = Project_Tradesheet.cell_value(r+2,14+R)

if ScriptConfig['Increased reuse'] == 'Moderate':      
  for p in range(10,16):
       for r in range(0,Par_NoOfRegions):
          for R in range(0,Par_NoOfRegions):
              if r == R:
                   Par_Chi_Reuse_inf[p,r,R] = Par_Chi_Reuse_inf[p,r,R] + 0.05
                   Par_Chi_Reuse[p,r,R] = Par_Chi_Reuse[p,r,R] + 0.05
                   
                   if  Par_Chi_Reuse_inf[p,r,R]> 1:
                        Par_Chi_Reuse_inf[p,r,R] = 1
                        
                   if Par_Chi_Reuse[p,r,R] > 1:
                       Par_Chi_Reuse[p,r,R] = 1

if ScriptConfig['Increased reuse'] == 'Ambitious':      
  for p in range(10,16):
       for r in range(0,Par_NoOfRegions):
          for R in range(0,Par_NoOfRegions):
              if r == R:
                  Par_Chi_Reuse_inf[p,r,R] = Par_Chi_Reuse_inf[p,r,R]+ 0.20
                  Par_Chi_Reuse[p,r,R] = Par_Chi_Reuse[p,r,R] + 0.20
                  
                  if  Par_Chi_Reuse_inf[p,r,R]> 1:
                        Par_Chi_Reuse_inf[p,r,R] = 1
                        
                  if Par_Chi_Reuse[p,r,R] > 1:
                       Par_Chi_Reuse[p,r,R] = 1
                       
                  
Par_Psi_EoL_Trade = np.zeros((Par_NoOfProducts,Par_NoOfRegions,Par_NoOfRegions)) 
if Par_NoOfRegions==1:
    for p in range(0,Par_NoOfProducts):
        for r in range(0,Par_NoOfRegions):
            for R in range(0,Par_NoOfRegions):
                Par_Psi_EoL_Trade[p,r,R] = 1   
else:         
    for p in range(0,Par_NoOfProducts):
        for r in range(0,Par_NoOfRegions):
            for R in range(0,Par_NoOfRegions):
                      if r==R:
                          Par_Psi_EoL_Trade[p,r,R] = 1
                      else:
                              Par_Psi_EoL_Trade[p,r,R] = 0

    for p in range(12,16):
        for r in range(0,Par_NoOfRegions):
            for R in range(0,Par_NoOfRegions):
                Par_Psi_EoL_Trade[p,r,R] = Project_Tradesheet.cell_value(r+13,2+R)



Par_Psi_EoL_Trade_inf = np.zeros((Par_NoOfProducts,Par_NoOfRegions,Par_NoOfRegions)) 
if Par_NoOfRegions==1:
    for p in range(0,Par_NoOfProducts):
        for r in range(0,Par_NoOfRegions):
            for R in range(0,Par_NoOfRegions):
                Par_Psi_EoL_Trade_inf[p,r,R] = 1   
else:         
    for p in range(0,Par_NoOfProducts):
        for r in range(0,Par_NoOfRegions):
            for R in range(0,Par_NoOfRegions):
                      if r==R:
                          Par_Psi_EoL_Trade_inf[p,r,R] = 1
                      else:
                              Par_Psi_EoL_Trade_inf[p,r,R] = 0


                

if ScriptConfig['No informal trade '] == 'no':   
    for p in range(12,16):
        for r in range(0,Par_NoOfRegions):
            for R in range(0,Par_NoOfRegions):
                Par_Psi_EoL_Trade_inf[p,r,R] = Project_Tradesheet.cell_value(r+13,14+R)
                
#%% Define System variables
      
Mylog.info('Define system variables.<br>')
#Define Process and market flows
# Define external input vector F_0_IV(t,r,p):
F_0_IV  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define final consumption vector F_IV_A(t,r,p):
F_IV_A  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define flow of copper in EoL products F_y(t,r,p) (internal flow):
F_y    = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define flow to domestic EoL products treatment in registered recycling practices, F_A_I(t,r,p):
F_A_I  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define flow to domestic EoL products treatment in informal recycling practices, F_A_I(t,r,p):
F_A_Iinf  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define flow EoL products to domestic scrap treatment, in registered recycling practices, F_I_B(t,r,p):
F_I_B  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts)) 
# Define flow EoL products to domestic scrap treatment, in informal recycling practices, F_I_Binf(t,r,p):
F_I_Binf= np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts)) 
# Define flow of products for re-use, to be inserted into stocks domestically or exported to other regions in registered re-use practices, F_I_IV(t,r,p):
F_I_IV  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define flow of products for re-use, to be inserted into stocks domestically or exported to other regions in informal re-use practices, F_I_IVinf(t,r,p):
F_I_IVinf= np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define flow of (old) scrap from EoL treatment to scrap market in registered scrap treatment practices, F_B_II(t,r,s):
F_B_II  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfScraps)) 
# Define flow of (old) scrap from EoL treatment to scrap market in informal scrap treatment practices, F_B_IIinf(t,r):
F_B_IIinf = np.zeros((Par_NoOfYears,Par_NoOfRegions)) 
#Define flow of total scrap material for copper recovery in registered recycling practices, F_II_C(t,r,o):
F_II_C  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfRecyclingRoutes))
#Define flow of total scrap material for copper recovery in informal recycling practices, F_II_Cinf(t,r,o):
F_II_Cinf= np.zeros((Par_NoOfYears,Par_NoOfRegions))
#Define flow of total remelted material, to secondary metal markets in registered recycling practices, F_C_III(t,r,c):
F_C_III  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfSecMetals)) 
#Define flow of total remelted material, to secondary metal markets in informal recycling practices, F_C_III(t,r):
F_C_IIIinf = np.zeros((Par_NoOfYears,Par_NoOfRegions)) 
#Define flow of total secondary material consumed from registered recovery processes, F_III_D(t,r,c):
F_III_D  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfSecMetals)) 
#Define flow of total secondary material consumed from informal recovery processes, F_III_Dinf(t,r):
F_III_Dinf = np.zeros((Par_NoOfYears,Par_NoOfRegions)) 
# Define flow of copper in manufactured recycled products to final product trade coming from registered wastemngm processes, F_D_IV(t,r,p):
F_D_IV  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts)) 
# Define flow of copper in manufactured recycled products to final product trade coming from informal wastemngm processes, F_D_IVinf(t,r,p):
F_D_IVinf  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts)) 
# Define fabrication scrap flow, F_D_II(t,r,s)
F_D_II  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfScraps)) 

 
#Define flows to the environment at different life stages 
#Obsolete stocks F_A_Env_Omega(t,r,p):
F_A_Env_Omega = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#FDissipative losses F_A_Env_Sigma(t,r,p):
F_A_Env_Sigma = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#Products which reach EoL but are not collected F_A_Env_Gamma(t,r,p)
F_A_Env_Gamma = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#Copper losses in scrap separation and sorting in registered recycling practices F_B_Env_Phi(t,r,s)
F_B_Env_Phi     = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfScraps))
#Copper losses in scrap separation and sorting in informal recycling practices F_B_Env_Phiinf(t,r)
F_B_Env_Phi_inf = np.zeros((Par_NoOfYears,Par_NoOfRegions))
#Copper losses in copper recovery in registered recycling practices F_C_Env_Theta(t,r,o)
F_C_Env_Theta   = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfRecyclingRoutes))
#Copper losses in copper recovery in informal recycling practices F_C_Env_Thetainf(t,r)
F_C_Env_Theta_inf= np.zeros((Par_NoOfYears,Par_NoOfRegions))

# Define stocks
# In use stock
S_1  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#Losses of products which reach EoL and are not collected
S_Env_Gamma= np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Obsolete stock
S_Env_Omega = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#Dissipative losses
S_Env_Sigma= np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#Losses from scrap separating and sorting
S_Env_Phi = np.zeros((Par_NoOfYears,Par_NoOfRegions))
#Losses from copper recovery
S_Env_Theta= np.zeros((Par_NoOfYears,Par_NoOfRegions))

# Total amount of copper in the system
S_Tot = np.zeros(Par_NoOfYears)


# Define Balance of use phase
Bal_A  = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
# Define Balance of Waste management industries:
Bal_B  = np.zeros((Par_NoOfYears,Par_NoOfRegions))
# Define Balance of Waste management industries in the informal recycling sector:
Bal_Binf= np.zeros((Par_NoOfYears,Par_NoOfRegions))
# Define Balance of the remelting industries
Bal_C = np.zeros((Par_NoOfYears,1))
# Define Balance of the fabrication sectors:
Bal_D = np.zeros((Par_NoOfYears,Par_NoOfRegions))

# Define market balances
#Define Balance of Final products market
Bal_IV    = np.zeros((Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts))
#Define Balance of EoL products market
Bal_I    = np.zeros((Par_NoOfYears,Par_NoOfProducts))
#Define balance of scrap markets
Bal_II    = np.zeros((Par_NoOfYears,1))
#Define balance of material markets
Bal_III    = np.zeros((Par_NoOfYears,1))
#Define system-wide mass balance:
Bal_System           = np.zeros(Par_NoOfYears)


#Define total Losses
Total_Losses_regions = np.zeros((Par_NoOfRegions))  
Total_Losses_share   = np.zeros((Par_NoOfRegions))  
Total_Losses_Omega   = np.zeros((Par_NoOfRegions)) 
Total_Losses_Sigma   = np.zeros((Par_NoOfRegions)) 
Total_Losses_Phi     = np.zeros((Par_NoOfRegions)) 
Total_Losses_Theta   = np.zeros((Par_NoOfRegions)) 
Total_Losses_Gamma   = np.zeros((Par_NoOfRegions)) 

EoLRR=np.zeros((Par_NoOfYears,Par_NoOfRegions))     
EoLRR_global=np.zeros((Par_NoOfYears)) 

#%% Peform caclulations
Mylog.info('<p><b>Performing model calculations.</b></p>')
SY = ScriptConfig['StartYear no.'] -2015 # index for start year
EY = ScriptConfig['Time horizon']  -2015 # index for end year
Mylog.info('<p>Modus: Trace fate of copper </p>')
Mylog.info('<p>Tracing fate of Copper  ' + ' starting in year ' + str(Par_Time[SY]) + ', ending in year ' + str(Par_Time[EY]) + '.</p>')

#F_0_IV, F_y,F_A_I , F_A_Iinf ,F_I_B,F_I_Binf,F_I_IV,F_I_IVinf,F_B_IIinf, F_II_C,
        #    F_II_Cinf, F_B_II,F_C_III, F_C_IIIinf, F_III_D, F_III_Dinf, F_D_IV, F_D_II, F_D_IVinf,
        #    F_C_Env_Theta_inf, F_C_Env_Theta, F_B_Env_Phi_inf, F_B_Env_Phi, F_A_Env_Gamma, 
            #
        #    F_A_Env_Sigma, F_A_Env_Omega, 
            
# perform a year-by-year computation of the outflow, recycling, and inflow
Mylog.info('<p>MaTrace Copper was successfully initialized. Starting to calculate the future material distribution.</p>')

def Matrace(Par_A_EolToScrap_Copper, Par_B_ScrapToRemeltingRoute, Par_C_RemeltingToSecondaryMetal, Par_D_AllocationCopperToProducts,
            Par_Sigma_Losses,Par_Omega_ObsoleteStocks, Par_Gamma_EoL_Collection_Rate_Copper, Par_inf_collection_rate, Par_Chi_Reuse, Par_Chi_Reuse_inf,
            Par_Psi_EoL_Trade_inf, Par_Phi_Scrap_Sorting_Efficiency, MaTrace_pdf,Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf,Par_Phi_inf_Scrap_Sorting_Efficiency,
            Par_Input_F_0_8,SY, EY):
    
    
    F_0_IV[SY,:,:] = Par_Input_F_0_8[SY,:,:]


    for CY in range(0,EY+1): # CY stands for 'current year'
        
        Mylog.info('<p>Performing calculations for year ' + str(Par_Time[CY]) + '.</p>')
        F_IV_A[SY,:,:] = F_0_IV[SY,:,:]
        
        Mylog.info('Step 1: Determine flow of copper in EoL products .<br>')
        for r in range(0,Par_NoOfRegions):
            for p in range(0,Par_NoOfProducts):
                # Use MaTrace_pdf to determine convolution of historic inflow with lifetime distribution
                F_y[CY,r,p] = (F_IV_A[0:CY,r,p] * MaTrace_pdf[CY,0:CY,r,p]).sum()       

        Mylog.info('Step 2:Obsolete stocks, dissipative losses and non-collected Eol products and flow of collected EoL products.<br>')
        # Determine losses due to dissipation
        F_A_Env_Sigma[CY,:,:]=np.einsum('ij,ij->ij',F_y[CY,:,:],Par_Sigma_Losses[:,:])
        # Determine obsolete stocks:
        F_A_Env_Omega[CY,:,:]=np.einsum('rp,rp->rp',F_y[CY,:,:]-F_A_Env_Sigma[CY,:,:],Par_Omega_ObsoleteStocks[:,:])
        #Determine Flow of EoL products collected for recovery from registered collection practices
        F_A_I[CY,:,:]= np.einsum('ij,ij->ij',(F_y[CY,:,:]-F_A_Env_Omega[CY,:,:]-F_A_Env_Sigma[CY,:,:]),(Par_Gamma_EoL_Collection_Rate_Copper[:,:]))
        #Determine Flow of EoL products collected for recovery from informal collection practices
        F_A_Iinf[CY,:,:]=np.einsum('ij,ij->ij',Par_inf_collection_rate[:,:], F_y[CY,:,:]-F_A_Env_Omega[CY,:,:]-F_A_Env_Sigma[CY,:,:])
        #Determine Flow of EoL products which reach EoL but are not collected 
        F_A_Env_Gamma[CY,:,:]=F_y[CY,:,:]-F_A_Iinf[CY,:,:] - F_A_I[CY,:,:] - F_A_Env_Omega[CY,:,:] - F_A_Env_Sigma[CY,:,:]
      
        Mylog.info('Step 3:Products for re-use and material sent to the waste management industries.<br>')
        # Determine flow of products for re-use in registered re-use enterprises:
        F_I_IV[CY,:,:]  = np.einsum('ij,jik->kj',F_A_I[CY,:,:],Par_Chi_Reuse[:,:,:])
        # Determine inflormal flow of products for re-use:
        F_I_IVinf[CY,:,:]  = np.einsum('ij,jik->kj',F_A_Iinf[CY,:,:],Par_Chi_Reuse_inf[:,:,:])
        # Determine flow to waste separation and sorting including export for waste management in registered practices :
        F_I_B[CY,:,:]  = np.einsum('ij,jik->kj', np.einsum('ij,ji->ij',F_A_I[CY,:,:],(1-Par_Chi_Reuse[:,:,:].sum(axis=2))), Par_Psi_EoL_Trade[:,:,:]) 
        # Determine flow to waste separation and sorting including export for waste management in informal practices :
        F_I_Binf[CY,:,:]  = np.einsum('ij,jik->kj', np.einsum('ij,ji->ij',F_A_Iinf[CY,:,:],(1-Par_Chi_Reuse_inf[:,:,:].sum(axis=2))), Par_Psi_EoL_Trade_inf[:,:,:]) 
        

        Mylog.info('Step 4: Scrap separation and sorting and lossed to landfills.<br>')
        # Determine flow of old scrap from EoL treatment to scrap market in registered recycing practices 
        F_B_II[CY,:,:] = np.einsum('...ij,...i', Par_A_EolToScrap_Copper[:,:,:],F_I_B[CY,:,:])*Par_Phi_Scrap_Sorting_Efficiency[:,:]
        # Determine flow of old scrap from EoL treatment to scrap market in informal recycing practices 
        F_B_IIinf[CY,:] = F_I_Binf[CY,:,:].sum(axis=1)* Par_Phi_inf_Scrap_Sorting_Efficiency[:]
        # Determine flow of losses to landfills in registered scrap separation and sorting practices
        F_B_Env_Phi[CY,:,:] = (1 - Par_Phi_Scrap_Sorting_Efficiency[:,:])*np.einsum('...ij,...i', Par_A_EolToScrap_Copper[:,:,:],F_I_B[CY,:,:])
        # Determine flow of losses to landfills in informal scrap separation and sorting practices
        F_B_Env_Phi_inf[CY,:] = (1 - Par_Phi_inf_Scrap_Sorting_Efficiency[:])*F_I_Binf[CY,:,:].sum(axis=1)

        Mylog.info('Step 5: Copper recovery and losses to landfills.<br>')
        # Flow to copper recovery processes in registered recycling practices
        F_II_C[CY,:,:]=(np.einsum('rst,rst->rt',Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[:,:,:],np.einsum('ijk,ij->ijk',Par_B_ScrapToRemeltingRoute[:,:,:],(F_B_II[CY,:,:]+F_D_II[CY-1,:,:]))))
        # Flow to copper recovery processes in informal recycling practices
        F_II_Cinf[CY,:]= Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[:]* (F_B_IIinf[CY,:])       
        #Losses from copper recovery processes in registered recycling practices
        F_C_Env_Theta[CY,:,:]=np.einsum('rst,rst->rt',(1-Par_Theta_Copper_recovery_from_scrap_in_recyclingroute[:,:,:]),np.einsum('ijk,ij->ijk',Par_B_ScrapToRemeltingRoute[:,:,:],(F_B_II[CY,:,:]+F_D_II[CY-1,:,:])))
        #Losses from copper recovery processes in informal recycling practices
        F_C_Env_Theta_inf[CY,:]= (1-Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf[:])* (F_B_IIinf[CY,:])
        # Flow to secondary metal markets in registered recycling practices
        F_C_III[CY,:,:]= np.einsum('...i,...ij->...j',F_II_C[CY,:,:],Par_C_RemeltingToSecondaryMetal[:,:,:])
        # Flow to secondary metal markets in informal recycling practices
        F_C_IIIinf[CY,:]=F_II_Cinf[CY,:]          
    
        Mylog.info('Step 6: Manufacturing, and re-distribution into products.<br>')
        # Determine flow of secondary material to manufacturing processes from registered recycling practices
        F_III_D[CY,:,:]  = F_C_III[CY,:,:]
        # Determine flow of secondary material to manufacturing processes from informal recycling practices
        F_III_Dinf[CY,:]=F_C_IIIinf[CY,:]
        # Determine net production of goods from recycled material from infromal recycling:
        F_D_IVinf[CY,:,:]=np.einsum('rp,r->rp', Par_D_AllocationCopperToProducts[:,:],(F_III_Dinf[CY,:]))
        # Determine net production of goods from recycled material from registered recycling practices:
        F_D_IV[CY,:,:]  = np.einsum('rp,pr->rp',np.einsum('rp,rc->rp', Par_D_AllocationCopperToProducts[:,:],(F_III_D[CY,:,:])),Par_Lambda_Fabrication_Efficiency[:,:])
        # Determine new scrap (fabrication scrap)
        F_D_II[CY,:,6]=np.einsum('rp,pr->rp',np.einsum('rp,rc->rp', Par_D_AllocationCopperToProducts[:,:],(F_C_III[CY,:,:])),(1-Par_Lambda_Fabrication_Efficiency[:,:])).sum(axis=1)
     
 
        Mylog.info('Step 7: Re-insert recycled and re-used goods into the stock.<br>')
        # Determine the total consumption of products as sum of external input from registered and informal recycling from re-used products from all regions, and recycled products from all regions
        F_IV_A[CY,:,:] = (F_D_IV[CY,:,:]+F_I_IV[CY,:,:]+F_0_IV[CY,:,:]+F_I_IVinf[CY,:,:]+ F_D_IVinf[CY,:,:])

        
        Mylog.info('Step 8: Determine system balance.<br>')    
        Bal_A[CY,:,:] =   F_A_I[CY,:,:] +F_A_Env_Sigma[CY,:,:]+F_A_Env_Omega[CY,:,:]+ F_A_Env_Gamma[CY,:,:] - F_y[CY,:,:]+F_A_Iinf[CY,:,:]
        Bal_B[CY,:] = F_I_B[CY,:,:].sum(axis = 1) + F_I_Binf[CY,:,:].sum(axis=1) - F_B_II[CY,:,:].sum(axis = 1)- F_B_IIinf[CY,:]  - F_B_Env_Phi[CY,:,:].sum(axis = 1) - F_B_Env_Phi_inf[CY,:]    
        Bal_Binf[CY,:] = F_I_Binf[CY,:,:].sum(axis = 1) - F_B_IIinf[CY,:] - F_B_Env_Phi_inf[CY,:] 
        Bal_C[CY,:]  = F_II_C[CY,:,:].sum() - F_C_III[CY,:,:].sum()
        Bal_D[CY,:]  = F_III_D[CY,:,:].sum(axis=1) - F_D_IV[CY,:,:].sum(axis=1) - F_D_II[CY,:,:].sum(axis=1) + F_III_Dinf[CY,:] - F_D_IVinf[CY,:,:].sum(axis=1)
        Bal_I[CY,:] = F_A_I[CY,:,:].sum() - F_I_IV[CY,:,:].sum() - F_I_B[CY,:,:].sum() + F_A_Iinf[CY,:,:].sum() - F_I_IVinf[CY,:,:].sum() - F_I_Binf[CY,:,:].sum()       
        Bal_II[CY,:]  =  F_D_II[CY-1,:].sum() + F_B_II[CY,:,:].sum() + F_B_IIinf[CY,:].sum() - F_II_C[CY,:,:].sum() -  F_II_Cinf[CY,:].sum()- F_C_Env_Theta_inf[CY,:].sum() - F_C_Env_Theta[CY,:,:].sum()       
        Bal_III[CY,:]  = F_C_III[CY,:,:].sum() - F_III_D[CY,:,:].sum()
        Bal_IV[CY,:,:]   = F_IV_A[CY,:,:] - F_0_IV[CY,:,:] -  F_I_IV[CY,:,:] - F_D_IV[CY,:,:] - F_I_IVinf[CY,:,:] -  F_D_IVinf[CY,:,:]
        
        Mylog.info('Step 9: Determine stocks.<br>')    

        if CY == 0:
            Mylog.info('Stock determination for year 0.<br>')
            S_1[0,:,:]                  = F_IV_A[0,:,:]    -F_y[0,:,:]    
            S_Env_Omega[0,:,:]          = F_A_Env_Omega[0,:,:]
            S_Env_Sigma[0,:,:]          = F_A_Env_Sigma[0,:,:]
            S_Env_Phi[0,:]            = F_B_Env_Phi[0,:,:].sum(axis=1) + F_B_Env_Phi_inf[0,:]
            S_Env_Theta[0,:]          = F_C_Env_Theta[0,:,:].sum(axis=1) + F_C_Env_Theta_inf[0,:]
            S_Env_Gamma[0,:,:]              = F_A_Env_Gamma[0,:,:]
            
        else:
            Mylog.info('Stock determination.<br>')
            S_1[CY,:,:]                 = S_1[CY-1,:,:]         + F_IV_A[CY,:,:] - F_y[CY,:,:]          
            S_Env_Omega[CY,:,:]         = S_Env_Omega[CY-1,:,:] + F_A_Env_Omega[CY,:,:]
            S_Env_Sigma[CY,:,:]         = S_Env_Sigma[CY-1,:,:] + F_A_Env_Sigma[CY,:,:]
            S_Env_Phi[CY,:]           = S_Env_Phi[CY-1,:]  + F_B_Env_Phi[CY,:,:].sum(axis=1) + F_B_Env_Phi_inf[CY,:]
            S_Env_Theta[CY,:]         = S_Env_Theta[CY-1,:] + F_C_Env_Theta[CY,:,:].sum(axis=1) + F_C_Env_Theta_inf[CY,:]  
            S_Env_Gamma[CY,:,:]             = S_Env_Gamma[CY-1,:,:]     + F_A_Env_Gamma[CY,:,:]
            
    return S_1, S_Env_Gamma, S_Env_Omega, S_Env_Sigma, S_Env_Phi, S_Env_Theta, S_Tot, Bal_A, Bal_B, Bal_Binf, Bal_C, Bal_D, Bal_I, Bal_II, Bal_III, Bal_IV, Bal_System, F_IV_A
          


S_1, S_Env_Gamma, S_Env_Omega, S_Env_Sigma, S_Env_Phi, S_Env_Theta, S_Tot, Bal_A, Bal_B, Bal_Binf, Bal_C, Bal_D, Bal_I, Bal_II, Bal_III, Bal_IV, Bal_System, F_IV_A = Matrace(Par_A_EolToScrap_Copper, Par_B_ScrapToRemeltingRoute, Par_C_RemeltingToSecondaryMetal, Par_D_AllocationCopperToProducts,
        Par_Sigma_Losses,Par_Omega_ObsoleteStocks, Par_Gamma_EoL_Collection_Rate_Copper, Par_inf_collection_rate, Par_Chi_Reuse, Par_Chi_Reuse_inf,
        Par_Psi_EoL_Trade_inf, Par_Phi_Scrap_Sorting_Efficiency, MaTrace_pdf,Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf,Par_Phi_inf_Scrap_Sorting_Efficiency,
        Par_Input_F_0_8,SY, EY)

##%%
###%% This section is for the MCS
###Define parameters according to uncertainty for MCS
#
#Par_sigma_D_AllocationCopperToProducts = np.zeros((Par_NoOfRegions,Par_NoOfProducts))
#A=np.zeros((Par_NoOfRegions,Par_NoOfProducts))
#B=np.zeros((Par_NoOfRegions,Par_NoOfProducts))
#regions = np.arange(0,Par_NoOfRegions)
#
#for R in range(0,Par_NoOfRegions):
#    Par_sigma_D_AllocationCopperToProducts[R,:] = Par_D_AllocationCopperToProducts[R,:]*0.3
#    #Now here we calculate the a and b paramters for the truncated normal distribution
#    A[R,:] = np.divide((0-Par_D_AllocationCopperToProducts[R,:]),Par_sigma_D_AllocationCopperToProducts[R,:], where=Par_D_AllocationCopperToProducts[R,:]!=0) #see documentation of scipy truncnorm
#    B[R,:] = np.divide((1-Par_D_AllocationCopperToProducts[R,:]),Par_sigma_D_AllocationCopperToProducts[R,:], where=Par_D_AllocationCopperToProducts[R,:]!=0)
#    #Now use list comprehension to build a list of truncnorm distributions for the different parameters. 
#Par_D_dict = {'Par_D_Uncertain_R{}'.format(R) : [scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(A[R,:],B[R,:],Par_D_AllocationCopperToProducts[R,:], Par_sigma_D_AllocationCopperToProducts[R,:])] for R in regions}
#    #Par_D_Uncertain.append([scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(A,B,Par_D_AllocationCopperToProducts[R,:], Par_sigma_D_AllocationCopperToProducts)])
#
#    
#x = np.linspace(-.1,1,100) #set the range
#fig = plt.figure(figsize=(20,20))#make a figure
#axes = fig.subplots(4,4)
#for i,(par,ax,mean) in enumerate(zip(Par_D_dict['Par_D_Uncertain_R8'][:-1],axes.flatten(),Par_D_AllocationCopperToProducts[8,:])):#the last one is zero so we don't plot it(Check this!!)
#    ax.plot(x,par.pdf(x),label='PDF')
#    ax.vlines(mean,ymin=0,ymax=par.pdf(x).max(),label='True mean')
#    ax.set_xlabel('Sector split $X$')
#    ax.set_ylabel('$P(X)$')
#    ax.set_title('Split distriution for {}'.format(Def_ProductNames[i]))
#    if i==0:
#        ax.legend(loc=1)
#fig.tight_layout()
#
##%% This section is for the MCS
##Define parameters according to uncertainty for MCS
#
#
#sample_size = 1000
#
#Par_A_EolToScrap_Copper_sample = np.zeros((sample_size,9,16,7))
#Par_B_ScrapToRemeltingRoute_sample = np.zeros((sample_size,9,7,2))
#Par_C_RemeltingToSecondaryMetal_sample = np.zeros((sample_size,9,2,1))
#Par_D_AllocationCopperToProducts_sample = np.zeros((sample_size,9,16))
#
#Par_Sigma_Losses_sample = np.zeros((sample_size,9,16)) #r,p
#Par_Omega_ObsoleteStocks_sample = np.zeros((sample_size,9,16))
#Par_Gamma_EoL_Collection_Rate_Copper_sample = np.zeros((sample_size,9,16))
#Par_Chi_Reuse_sample = np.zeros((sample_size,16,9,9))
#Par_Chi_Reuse_inf_sample = np.zeros((sample_size,16,9,9))
#Par_Psi_EoL_Trade_inf_sample = np.zeros((sample_size,16,9,9))
#Par_Phi_Scrap_Sorting_Efficiency_sample = np.zeros((sample_size,9,7))
#Par_Tau_sample = np.zeros((sample_size,9,16))
#Par_Phi_inf_Scrap_Sorting_Efficiency_sample = np.zeros((sample_size,9))
#
#### Defining standard deviations
## lu (low uncertainty), mu (medium uncertainty), hu (high uncertainty)
#lu = 0.1
#mu = 0.2
#hu = 0.3
#
## creating samples according to a truncated normal distribution for paramters which are not correlated (efficiencies and trade parameters) 
##(Par_Sigma_Losses _sample,Par_Phi_Scrap_Sorting_Efficiency_sample,Par_Omega_ObsoleteStocks_sample,Par_Gamma_EoL_Collection_Rate_Copper_sample,Par_Chi_Reuse_sample,Par_Chi_Reuse_inf_sample,Par_Psi_EoL_Trade_inf_sample)
#
## dimension products
#
#Ap=np.zeros((Par_NoOfRegions,Par_NoOfProducts))
#Bp=np.zeros((Par_NoOfRegions,Par_NoOfProducts))
#regions = np.arange(0,Par_NoOfRegions)
#x = np.linspace(-.1,1,100) #set the range
#
#
#Par_sigma_p = Par_Omega_ObsoleteStocks[:,:]*lu
#Par_mean_p = Par_Omega_ObsoleteStocks[:,:]
#
#for R in range(0,Par_NoOfRegions):
#    #Now here we calculate the a and b paramters for the truncated normal distribution with lower limit 0 and upper limit 1
#    Ap[R,:] = np.divide((0-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0) #see documentation of scipy truncnorm
#    Bp[R,:] = np.divide((1-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0)
#    #Now use list comprehension to build a list of truncnorm distributions for the different parameters. 
#Par_D_dict = {'Par_D_Uncertain_R{}'.format(R) :[scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(Ap[R,:],Bp[R,:],Par_mean_p[R,:], Par_sigma_p[R,:])] for R in regions}
#    #Par_D_Uncertain.append([scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(A,B,Par_D_AllocationCopperToProducts[R,:], Par_sigma_D_AllocationCopperToProducts)])
#
#
#for R in range(0,Par_NoOfRegions):
#    for i,(par,mean) in enumerate(zip(Par_D_dict['Par_D_Uncertain_R{}'.format(R)],Par_mean_p[R,:])):     
#        if mean != 0:
#            Par_Omega_ObsoleteStocks_sample[:,R,i] = par.rvs(size=sample_size)
#        else:
#            Par_Omega_ObsoleteStocks_sample[:,R,i] = np.zeros(sample_size)
#
#
#
#Par_sigma_p = Par_Sigma_Losses[:,:]*lu
#Par_mean_p = Par_Sigma_Losses[:,:]
#
#for R in range(0,Par_NoOfRegions):
#    #Now here we calculate the a and b paramters for the truncated normal distribution with lower limit 0 and upper limit 1
#    Ap[R,:] = np.divide((0-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0) #see documentation of scipy truncnorm
#    Bp[R,:] = np.divide((1-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0)
#    #Now use list comprehension to build a list of truncnorm distributions for the different parameters. 
#
#Par_D_dict = {'Par_D_Uncertain_R{}'.format(R) :[scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(Ap[R,:],Bp[R,:],Par_mean_p[R,:], Par_sigma_p[R,:])] for R in regions}
#    #Par_D_Uncertain.append([scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(A,B,Par_D_AllocationCopperToProducts[R,:], Par_sigma_D_AllocationCopperToProducts)])
#
#for R in range(0,Par_NoOfRegions):
#    for i,(par,mean) in enumerate(zip(Par_D_dict['Par_D_Uncertain_R{}'.format(R)],Par_mean_p[R,:])):     
#        if mean != 0:
#            Par_Sigma_Losses_sample[:,R,i] = par.rvs(size=sample_size)
#        else:
#            Par_Sigma_Losses_sample[:,R,i] = np.zeros(sample_size)
#
#
#Par_sigma_p = Par_Tau[:,:]*lu
#Par_mean_p = Par_Tau[:,:]  
#
#for R in range(0,Par_NoOfRegions):
#    #Now here we calculate the a and b paramters for the truncated normal distribution with lower limit 0 and upper limit 1
#    Ap[R,:] = np.divide((0-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0) #see documentation of scipy truncnorm
#    Bp[R,:] = np.divide((np.inf-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0)
#    #Now use list comprehension to build a list of truncnorm distributions for the different parameters. 
#
#Par_D_dict = {'Par_D_Uncertain_R{}'.format(R) :[scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(Ap[R,:],Bp[R,:],Par_mean_p[R,:], Par_sigma_p[R,:])] for R in regions}
#    #Par_D_Uncertain.append([scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(A,B,Par_D_AllocationCopperToProducts[R,:], Par_sigma_D_AllocationCopperToProducts)])
#
#for R in range(0,Par_NoOfRegions):
#    for i,(par,mean) in enumerate(zip(Par_D_dict['Par_D_Uncertain_R{}'.format(R)],Par_mean_p[R,:])):     
#        if mean != 0:
#            Par_Tau_sample[:,R,i] = par.rvs(size=sample_size)
#        else:
#            Par_Tau_sample[:,R,i] = np.zeros(sample_size)
#
#
#
## dimension scraps
#
#    
#As=np.zeros((Par_NoOfRegions,Par_NoOfScraps))
#Bs=np.zeros((Par_NoOfRegions,Par_NoOfScraps))
#    
#Par_sigma_p = Par_Phi_Scrap_Sorting_Efficiency[:,:]*0.1
#Par_mean_p = Par_Phi_Scrap_Sorting_Efficiency[:,:]
#
#for R in range(0,Par_NoOfRegions):
#    #Now here we calculate the a and b paramters for the truncated normal distribution with lower limit 0 and upper limit 1
#    As[R,:] = np.divide((0-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0) #see documentation of scipy truncnorm
#    Bs[R,:] = np.divide((1-Par_mean_p[R,:]),Par_sigma_p[R,:], where=Par_mean_p[R,:]!=0)
#    #Now use list comprehension to build a list of truncnorm distributions for the different parameters. 
#Par_D_dict = {'Par_D_Uncertain_R{}'.format(R) :[scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(As[R,:],Bs[R,:],Par_mean_p[R,:], Par_sigma_p[R,:])] for R in regions}
#    #Par_D_Uncertain.append([scipy.stats.truncnorm(a= a,b=b,loc=mu,scale=sigma) for a,b,mu,sigma in zip(A,B,Par_D_AllocationCopperToProducts[R,:], Par_sigma_D_AllocationCopperToProducts)])
# 
#
#for R in range(0,Par_NoOfRegions):
#    for i,(par,mean) in enumerate(zip(Par_D_dict['Par_D_Uncertain_R{}'.format(R)],Par_mean_p[R,:])):     
#        if mean != 0:
#            Par_Phi_Scrap_Sorting_Efficiency_sample[:,R,i] = par.rvs(size=sample_size)
#        else:
#            Par_Phi_Scrap_Sorting_Efficiency_sample[:,R,i] = np.zeros(sample_size)
#            
#A=np.zeros((Par_NoOfRegions))
#B=np.zeros((Par_NoOfRegions))
#
#A[:]=np.zeros((Par_NoOfRegions))
#B[:]=np.zeros((Par_NoOfRegions))
#
#for R in range(0,Par_NoOfRegions):
#    #Now here we calculate the a and b paramters for the truncated normal distribution with lower limit 0 and upper limit 1
#    A[R] = (0-(Par_Phi_inf_Scrap_Sorting_Efficiency[R]))/(Par_Phi_inf_Scrap_Sorting_Efficiency[R]*hu) #see documentation of scipy truncnorm
#    B[R] = (1-(Par_Phi_inf_Scrap_Sorting_Efficiency[R]))/(Par_Phi_inf_Scrap_Sorting_Efficiency[R]*hu)
#    #Now use list comprehension to build a list of truncnorm distributions for the different parameters. 
#    Par_Phi_inf_Scrap_Sorting_Efficiency_sample[:,R] =scipy.stats.truncnorm(a= A[R],b=B[R],loc=(Par_Phi_inf_Scrap_Sorting_Efficiency[R]),scale=(Par_Phi_inf_Scrap_Sorting_Efficiency[R]*hu)).rvs(size=sample_size)
#
#            
##%%
## creating samples for paramters in which values are correlated ( process parameters with sum = 1)   
##Par_std_A_EolToScrap_Copper_sample, Par_std_B_ScrapToRecyclingRoute_sample ,Par_std_D_AllocationCopperToProducts_sample    
#    
#def ell (alpha, mean, std):
#    #alpha, mean, std = theta
#    return   ((scipy.stats.dirichlet(alpha).mean()-mean)**2).sum()+\
#                ((np.sqrt(scipy.stats.dirichlet(alpha).var())-std**2)**2).sum()
#                
#for R in range(0,Par_NoOfRegions):
#    for p in range (0,Par_NoOfProducts):
#        expt_value = Par_A_EolToScrap_Copper[R,p,:] + 1E-9  #cannot be zero
#        sigma = Par_A_EolToScrap_Copper[R,p,:] *0.1
#        
#        alpha_opt = scipy.optimize.minimize(ell, x0=expt_value, args=(expt_value,
#                            sigma),bounds=[(1e-9,np.inf) for i in range(len(expt_value))])
#
#        opt_model = scipy.stats.dirichlet(alpha=alpha_opt['x'])
#        Par_A_EolToScrap_Copper_sample[:,R,p,:] = opt_model.rvs(size=sample_size) #each sample has shape of expt_value and sum 1
#
#
#for R in range(0,Par_NoOfRegions):
#    for s in range (0,Par_NoOfScraps):
#        expt_value = Par_B_ScrapToRemeltingRoute[R,s,:] + 1E-9
#        sigma = Par_B_ScrapToRemeltingRoute[R,s,:]*0.1
#
#        alpha_opt = scipy.optimize.minimize(ell, x0=expt_value, args=(expt_value,
#                            sigma),bounds=[(1e-9,np.inf) for i in range(len(expt_value))])
#
#        opt_model = scipy.stats.dirichlet(alpha=alpha_opt['x'])
#        Par_B_ScrapToRemeltingRoute_sample[:,R,s,:] = opt_model.rvs(size=sample_size) #each sample has shape of expt_value and sum 1
#
#for R in range(0,Par_NoOfRegions):
#        expt_value = Par_D_AllocationCopperToProducts[R,:] + 1E-9
#        sigma = Par_D_AllocationCopperToProducts[R,:] *0.1
#
#        alpha_opt = scipy.optimize.minimize(ell, x0=expt_value, args=(expt_value,
#                            sigma),bounds=[(1e-9,np.inf) for i in range(len(expt_value))])
#
#        opt_model = scipy.stats.dirichlet(alpha=alpha_opt['x'])
#        Par_D_AllocationCopperToProducts_sample[:,R,:] = opt_model.rvs(size=sample_size) #each sample has shape of expt_value and sum 1     
#          
#Collectionrate = np.zeros((9,16,3))                     
#Collectionrate_sample=np.zeros((sample_size, 9,16,3))  
#Collectionrate[:,:,0] = Par_Gamma_EoL_Collection_Rate_Copper        
#Collectionrate[:,:,1]=Par_inf_collection_rate                    
#Collectionrate[:,:,2]=1- (Par_Gamma_EoL_Collection_Rate_Copper+Par_inf_collection_rate)
#
#for R in range(0,Par_NoOfRegions):
#    for i in range(0,Par_NoOfProducts):
#        expt_value = Collectionrate[R,i,:] + 1E-9
#        sigma = Collectionrate[R,i,0] *0.1
#        sigma = Collectionrate[R,i,1] *0.3
#
#        alpha_opt = scipy.optimize.minimize(ell, x0=expt_value, args=(expt_value,
#                            sigma),bounds=[(1e-9,np.inf) for i in range(len(expt_value))])
#
#        opt_model = scipy.stats.dirichlet(alpha=alpha_opt['x'])
#        Collectionrate_sample[:,R,i,:] = opt_model.rvs(size=sample_size) #each sample has shape of expt_value and sum 1     
#
#
#
#              
#S_1_MCS = np.zeros((1000,Par_NoOfYears,9,16))
#F_IV_A_MCS = np.zeros((1000,Par_NoOfYears,9,16))
#        
#for i in range (0,1000): 
#    Par_A_EolToScrap_Copper = Par_A_EolToScrap_Copper_sample[i,:,:,:]
#    Par_B_ScrapToRemeltingRoute = Par_B_ScrapToRemeltingRoute_sample[i,:,:,:]
#    Par_D_AllocationCopperToProducts = Par_D_AllocationCopperToProducts_sample[i,:,:]
#    Par_Sigma_Losses = Par_Sigma_Losses_sample[i,:,:]
#    Par_Omega_ObsoleteStocks = Par_Omega_ObsoleteStocks_sample[i,:,:]
#    Par_Phi_Scrap_Sorting_Efficiency = Par_Phi_Scrap_Sorting_Efficiency_sample[i,:,:]
#    Par_Gamma_EoL_Collection_Rate_Copper=Collectionrate_sample[i,:,:,0]
#    Par_inf_collection_rate=Collectionrate_sample[i,:,:,1]
#    Par_Tau = Par_Tau_sample[i,:,:]
#    Par_Phi_inf_Scrap_Sorting_Efficiency = Par_Phi_inf_Scrap_Sorting_Efficiency_sample[i,:]
#    
#    MaTrace_pdf = np.zeros((Par_NoOfYears,Par_NoOfYears,Par_NoOfRegions,Par_NoOfProducts)) 
#    AgeMatrix   = np.zeros((Par_NoOfYears,Par_NoOfYears))
#    for c in range(0, Par_NoOfYears):  # cohort index
#        for y in range(c + 1, Par_NoOfYears):
#            AgeMatrix[y,c] = y-c
#    for R in range(0,Par_NoOfRegions):
#        for P in range(0,Par_NoOfProducts):
#            MaTrace_pdf[:,:,R,P] = scipy.stats.truncnorm((0-Par_Tau[R,P])/Par_Sigma[R,P], np.inf, loc=Par_Tau[R,P], scale=Par_Sigma[R,P]).pdf(AgeMatrix)  # Call scipy's Norm function with Mean, StdDev, and Age
#    # No discard in historic years and year 0:
#    for m in range(0,Par_NoOfYears):
#        MaTrace_pdf[0:m+1,m,:,:] = 0
#   
#    
#    S_1, S_Env_Gamma, S_Env_Omega, S_Env_Sigma, S_Env_Phi, S_Env_Theta, S_Tot, Bal_A, Bal_B, Bal_Binf, Bal_C, Bal_D, Bal_I, Bal_II, Bal_III, Bal_IV, Bal_System, F_IV_A = Matrace(Par_A_EolToScrap_Copper, Par_B_ScrapToRemeltingRoute, Par_C_RemeltingToSecondaryMetal, Par_D_AllocationCopperToProducts,
#            Par_Sigma_Losses,Par_Omega_ObsoleteStocks, Par_Gamma_EoL_Collection_Rate_Copper, Par_inf_collection_rate, Par_Chi_Reuse, Par_Chi_Reuse_inf,
#            Par_Psi_EoL_Trade_inf, Par_Phi_Scrap_Sorting_Efficiency, MaTrace_pdf,Par_Theta_Copper_recovery_from_scrap_in_recyclingroute_inf,Par_Phi_inf_Scrap_Sorting_Efficiency,
#            Par_Input_F_0_8,SY, EY)
#    
#    S_1_MCS[i,:,:,:] = S_1
#    
#    F_IV_A_MCS[i,:,:,:] = F_IV_A
#  
#Tot_S_samples =  S_1_MCS.sum(axis=2).sum(axis=2)
#Tot_S_025, Tot_S_16, Tot_S_median, Tot_S_84, Tot_S_975 = np.percentile(Tot_S_samples,[2.5,16,50,84,97.5],axis=0)
# 
#Tot_F_IV_A_samples =  F_IV_A_MCS.sum(axis=2).sum(axis=2)
#Tot_F_IV_A_025, Tot_F_IV_A_16, Tot_F_IV_A_median, Tot_F_IV_A_84, Tot_F_IV_A_975 = np.percentile(Tot_F_IV_A_samples,[2.5,16,50,84,97.5],axis=0)
#
#### Plot mcs stock plot
#Timex = np.arange(0,Par_NoOfYears,1)
#                
#MCS_Path = str(Path_Result)
#plt.xlabel('Year')
#plt.ylabel('Primary copper produced in 2015 [1000 metric tons]')
#x = list(range(0,86))
#Par_Time=list(range(2015,2101))
#
#plt.plot(Timex+2015,S_1.sum(axis=1).sum(axis=1), 'k')
#plt.fill_between(Timex+2015,Tot_S_975,Tot_S_025, alpha=0.3)
#plt.fill_between(Timex+2015,Tot_S_84,Tot_S_16, alpha=0.7)
#plt.plot(Timex+2015,Tot_S_median[Timex], 'k--')      
#plt.legend(['mean copper in use','median copper in use','2$\sigma$ confidence interval','1$\sigma$ confidence interval'])
#my_xticks = [Par_Time]
#xmin, xmax, ymin, ymax = 2015, 2100, 0, 20000
#plt.axis([xmin, xmax, ymin, ymax])
#plt.savefig(MCS_Path +'MCS_BAU_only.svg',acecolor='w',edgecolor='w', bbox_inches='tight')
#plt.savefig(MCS_Path +'MCS_BAU_only.png',acecolor='w',edgecolor='w', dpi=500)
#
#plt.show()
# 
##
#
#
#####  Calculation of longevity and corcularity ##
#### longevity: summing up the total in-use stock for every year and devide this by the initial inflow at year t=0
#longevity_median = round(Tot_S_median[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#longevity_975 = round(Tot_S_975[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#longevity_025 = round(Tot_S_025[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#longevity_84 = round(Tot_S_84[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#longevity_16 = round(Tot_S_16[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#
### Calculate circularity
#### To calculate the circularity we sum over the inflow in the use phsae and devide this by the initial inflow
#n_median = round(Tot_F_IV_A_median[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#n_975 = round(Tot_F_IV_A_975[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#n_025 = round(Tot_F_IV_A_025[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#n_84 = round(Tot_F_IV_A_84[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
#n_16 = round(Tot_F_IV_A_16[:].sum()/Par_Input_F_0_8[SY,:,:].sum(),2)
##
###

#%%
    

tau= round(S_1.sum()/19800,1)
n_circ = round(F_IV_A.sum()/Par_Input_F_0_8[SY,:,:].sum(),2)

print(Name_Scenario)
print('tau')
print(tau)
print('n')
print(n_circ)
a=35
    
Total_Losses_tot=S_Env_Omega[a,:,:].sum()+ S_Env_Sigma[a,:,:].sum()+ S_Env_Phi[a,:].sum()+ S_Env_Theta[a,:].sum() + S_Env_Gamma[a,:,:].sum()

Total_Losses_lifestages = [S_Env_Gamma[a,:,:].sum(),S_Env_Omega[a,:,:].sum(),S_Env_Phi[a,:].sum(),S_Env_Sigma[a,:,:].sum(),S_Env_Theta[a,:].sum()] 

for r in range(0,Par_NoOfRegions):
    Total_Losses_regions[r]=round(((S_Env_Omega[a,r,:].sum()+ S_Env_Sigma[a,r,:].sum()+ S_Env_Phi[a,r].sum()+ S_Env_Theta[a,r].sum() + S_Env_Gamma[a,r,:].sum())))

    Total_Losses_share[r]=Total_Losses_regions[r]/ Total_Losses_tot

    Total_Losses_Omega[r]   = S_Env_Omega[a,r,:].sum()
    Total_Losses_Sigma[r]   = S_Env_Sigma[a,r,:].sum()
    Total_Losses_Phi[r]     = S_Env_Phi[a,r].sum()
    Total_Losses_Theta[r]   = S_Env_Theta[a,r].sum()
    Total_Losses_Gamma[r]   = S_Env_Gamma[a,r,:].sum()
    
    Losses=[S_Env_Gamma[a,r,:].sum(),S_Env_Omega[a,r,:].sum(),S_Env_Phi[a,r].sum()+S_Env_Sigma[a,r,:].sum(),S_Env_Theta[a,r].sum()]
    
#            
    
        
    
GN=[3] # Western Europe
GN_Share_Losses_Omega = Total_Losses_Omega[GN].sum()
GN_Share_Losses_Sigma = Total_Losses_Sigma[GN].sum()
GN_Share_Losses_Phi = Total_Losses_Phi[GN].sum()  
GN_Share_Losses_Theta = Total_Losses_Theta[GN].sum()
GN_Share_Losses_Gamma = Total_Losses_Gamma[GN].sum()
GN_in_use = S_1[35,GN,:].sum()


GS=[0] # China
GS_Share_Losses_Omega = Total_Losses_Omega[GS].sum()
GS_Share_Losses_Sigma = Total_Losses_Sigma[GS].sum()
GS_Share_Losses_Phi = Total_Losses_Phi[GS].sum()
GS_Share_Losses_Theta = Total_Losses_Theta[GS].sum()
GS_Share_Losses_Gamma = Total_Losses_Gamma[GS].sum()
GS_in_use = S_1[35,GS,:].sum()

GL=[7] # Africa
GL_Share_Losses_Omega = Total_Losses_Omega[GL].sum()
GL_Share_Losses_Sigma = Total_Losses_Sigma[GL].sum()
GL_Share_Losses_Phi = Total_Losses_Phi[GL].sum()
GL_Share_Losses_Theta = Total_Losses_Theta[GL].sum()
GL_Share_Losses_Gamma = Total_Losses_Gamma[GL].sum()
GL_in_use = S_1[35,GL,:].sum()

GP=[2] # NAM
GP_Share_Losses_Omega = Total_Losses_Omega[GP].sum()
GP_Share_Losses_Sigma = Total_Losses_Sigma[GP].sum()
GP_Share_Losses_Phi = Total_Losses_Phi[GP].sum()
GP_Share_Losses_Theta = Total_Losses_Theta[GP].sum()
GP_Share_Losses_Gamma = Total_Losses_Gamma[GP].sum()
GP_in_use = S_1[35,GP,:].sum()



results_share_GN=[Name_Scenario, 'Western Europe', GN_Share_Losses_Omega, GN_Share_Losses_Sigma ,GN_Share_Losses_Phi,GN_Share_Losses_Theta,GN_Share_Losses_Gamma,GN_in_use]
results_share_GS=[Name_Scenario,'China', GS_Share_Losses_Omega, GS_Share_Losses_Sigma ,GS_Share_Losses_Phi,GS_Share_Losses_Theta,GS_Share_Losses_Gamma, GS_in_use]
results_share_GL=[Name_Scenario,'Africa', GL_Share_Losses_Omega, GL_Share_Losses_Sigma ,GL_Share_Losses_Phi,GL_Share_Losses_Theta,GL_Share_Losses_Gamma, GL_in_use]
results_share_GP=[Name_Scenario,'NAM', GP_Share_Losses_Omega, GP_Share_Losses_Sigma ,GP_Share_Losses_Phi,GP_Share_Losses_Theta,GP_Share_Losses_Gamma, GP_in_use]



results_regions = [Total_Losses_regions]
results_in_use = [S_1[35,:,:].sum()]
    
    
results_regions_GN_df = pd.DataFrame( results_share_GN)
results_regions_GS_df = pd.DataFrame( results_share_GS)
results_regions_GL_df = pd.DataFrame( results_share_GL)
results_regions_GP_df = pd.DataFrame( results_share_GP)

results_regions_df = pd.DataFrame(results_regions)
results_in_use_df = pd.DataFrame(results_in_use)

writer=pd.ExcelWriter(str(Project_MainPath) + '\\General_Results\\' +  'Pi_Chart_regions_results.xlsx',engine='openpyxl')
   
          
book=load_workbook(str(Project_MainPath) + '\\General_Results\\' + 'Pi_Chart_regions_results.xlsx')

writer=pd.ExcelWriter(str(Project_MainPath) + '\\General_Results\\' + 'Pi_Chart_regions_results.xlsx',engine='openpyxl')

writer.book=book

writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
   
results_regions_GN_df.to_excel(writer, sheet_name='Pi_charts',startcol=int(Number_Scenario),startrow=1,index=False)
results_regions_GS_df.to_excel(writer, sheet_name='Pi_charts',startcol=int(Number_Scenario),startrow=10,index=False) 
results_regions_GL_df.to_excel(writer, sheet_name='Pi_charts',startcol=int(Number_Scenario),startrow=19,index=False) 
results_regions_GP_df.to_excel(writer, sheet_name='Pi_charts',startcol=int(Number_Scenario),startrow=28,index=False) 

results_regions_df.to_excel(writer, sheet_name='Regional_losses',startcol=0,startrow=int(Number_Scenario)*2,index=False) 
results_in_use_df.to_excel(writer, sheet_name='Regional_losses',startcol=12,startrow=int(Number_Scenario)*2,index=False) 
writer.save()








Mylog.info('<p>Create plots.</p>')
        
#%% Plots       

if ScriptConfig['Time horizon'] == 2100 or 2300: 
    print('yes')
    if ScriptConfig['Modus'] == 'Trace 2015 mined copper':
                
                            
                
                fig = plt.figure()

 #               matplotlib.rc('font', size=10)


                plt.xlabel('Year')
                plt.ylabel('Copper mined in 2015 [kt]')
                x = list(range(0,86))
                Par_Time=list(range(2015,2101))
                colormap = plt.cm.gist_ncar
                plt.fill_between(Par_Time, 0, S_1[x,:,:].sum(axis=1).sum(axis=1) ,facecolor='#F5F5DC',label='Total copper stock')
                             

                #plot lines with sector stocks
                plt.plot(Par_Time, S_1[x,:,12:16].sum(axis=1).sum(axis=1),'b',label='Consumer & Electronics')
                plt.plot(Par_Time, S_1[x,:,0:5].sum(axis=1).sum(axis=1),'y',label='Building & Construction')
         #       plt.plot(Par_Time, S_1[x,:,0:5].sum(axis=1).sum(axis=1)-S_Brass[x,:,:].sum(axis=1).sum(axis=1),'--r')
                plt.plot(Par_Time, S_1[x,:,5:7].sum(axis=1).sum(axis=1),'g',label='Infrastructure')
                plt.plot(Par_Time, S_1[x,:,7:9].sum(axis=1).sum(axis=1),'r',label='Industrial')
                plt.plot(Par_Time, S_1[x,:,9:12].sum(axis=1).sum(axis=1),'c',label='Transport')
        
                #plot losses and fill between
                plt.fill_between(Par_Time, (S_1[x,:,:].sum(axis=1).sum(axis=1)), S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1),facecolor='whitesmoke',label='Obsolete stocks')
                plt.fill_between(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1), S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1),facecolor='lightgrey',label='Dissipative losses')
                plt.fill_between(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1), S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Phi[x,:].sum(axis=1),facecolor='darkgrey',label='Losses due to scrap and waste separation and sorting')
                plt.fill_between(Par_Time,  S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Phi[x,:].sum(axis=1), S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Phi[x,:].sum(axis=1)+S_Env_Theta[x,:].sum(axis=1),facecolor='dimgrey',label='Losses by copper recovery from scrap in recycling route')
                plt.fill_between(Par_Time,   S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Phi[x,:].sum(axis=1)+S_Env_Theta[x,:].sum(axis=1), S_1[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1)+S_Env_Phi[x,:].sum(axis=1)+S_Env_Theta[x,:].sum(axis=1)+ S_Env_Gamma[x,:,:].sum(axis=1).sum(axis=1) ,facecolor='k',label='Not collected copper items')
                  
        
                #define shape of figure
                my_xticks = [Par_Time]
             #   plt.legend(frameon=True)
             #   plt.legend(bbox_to_anchor=(1.04,1), loc="lower center")
                xmin, xmax, ymin, ymax = 2015, 2100, 0, 20000
                plt.axis([xmin, xmax, ymin, ymax])
                plt.savefig(str(Path_Result) +'stockplot.svg',acecolor='w',edgecolor='w', bbox_inches='tight')
                plt.savefig(str(Path_Result) +'stockplot.png',acecolor='w',edgecolor='w', dpi=500)
        

#
#                Figure_2_paper =pd.DataFrame({#'Year':Par_Time, \
#                                              Name_Scenario + '_Consumer & Electronics':S_1[x,:,12:16].sum(axis=1).sum(axis=1), \
#                                              Name_Scenario + '_Building & Construction':S_1[x,:,0:5].sum(axis=1).sum(axis=1), \
#                                              Name_Scenario + '_Infrastructure':S_1[x,:,5:7].sum(axis=1).sum(axis=1), \
#                                              Name_Scenario + '_Industrial':S_1[x,:,7:9].sum(axis=1).sum(axis=1), \
#                                              Name_Scenario + '_Transport':S_1[x,:,9:12].sum(axis=1).sum(axis=1), \
#                                              Name_Scenario + '_Obsolete stocks':S_Env_Omega[x,:,:].sum(axis=1).sum(axis=1) ,\
#                                              Name_Scenario + '_Dissipative losses':S_Env_Sigma[x,:,:].sum(axis=1).sum(axis=1), \
#                                              Name_Scenario + '_Losses due to scrap and waste separation and sorting':S_Env_Phi[x,:].sum(axis=1), \
#                                              Name_Scenario + '_Losses by copper recovery from scrap in recycling route':S_Env_Theta[x,:].sum(axis=1), \
#                                              Name_Scenario + '_Not collected copper items':S_Env_Gamma[x,:,:].sum(axis=1).sum(axis=1) 
#
#                                              })
#                                              
#                                              #,,,,,]
#                
#               # Figure_2_paper_df = pd.DataFrame(Figure_2_paper)
#                  
#
#                book=load_workbook('C:\\Users\\sklose\\Documents\\Writings\\Paper\\Paper_Number_1\\Submit_JIE\\supporting_information_1.xlsx')
#                
#                writer=pd.ExcelWriter('C:\\Users\\sklose\\Documents\\Writings\\Paper\\Paper_Number_1\\Submit_JIE\\supporting_information_1.xlsx',engine='openpyxl')
#                
#                writer.book=book
#                
#                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#                
#                
#                Figure_2_paper.to_excel(writer, sheet_name='figure 2',startcol=34,startrow=0,index=False)
#                
#                
#                writer.save()       
#
#                #calculate average lifetime of copper in the technosphere
#                Figurecounter = 2
#                plt.figure(2)
#                
#                x1=list(range(0,100))
#                y= S_1[x,:,:].sum(axis=1).sum(axis=1)
#                def func(x, a, b, c):
#                    return a*np.exp(-c*(x-b))
#                popt, pcov = curve_fit(func, x, y)
#                #plot(x,y)
#              #  plot(x1,func(x1,*popt))
#               # show()
#        
#                #integrate underneath function!!
#                area1 = trapz(y, dx=1)
#                # Compute the area using the composite Simpson's rule.
#                area2 = simps(y, dx=1)
##        
#                x3=list(range(85,100000))
#                area3 = simps(func(x3,*popt), dx=1)
#                area=round(area1+area3)
#                
#               
#                
#          
##                plt.annotate('Average life-time of copper ='+ str(area),
##                    xy=(1.5, 0), xytext=(1, 5),
##                    xycoords=('axes fraction', 'figure fraction'),
##                    textcoords='offset points',
##                    size=14, ha='right', va='bottom')
#                
#                #calculate halftime (year when half of the copper is left in the anthroposphere)
#                
#                
#                y100= S_1[0:86,:,:].sum()/19800
#                Circ=round(y100/85,2)
#                print(Circ)
#                Total_copper=F_0_IV[SY,:,:].sum()
#                Half_copper=Total_copper/2
#                
#                Halftime=str(min(m for m in range (0,Par_NoOfYears) if S_1[m,:,:].sum(axis=0).sum(axis=0) < Half_copper))
#                Halftime_year=int(Halftime) + 2015
#
#                #calculate Amount of copper left afer 50 years
#                Copper_50=round(S_1[50,:,:].sum(axis=0).sum(axis=0)/19800,2)
#               
#                #number of circles of a copper atom on average
#                n = round(F_IV_A[:,:,:].sum()/19800,2)
#                
#                n_regions =np.zeros((Par_NoOfRegions))
#                area_regions = np.zeros((Par_NoOfRegions))
#               
#                    
#                y_regions= S_1[:,:,:].sum(axis=1).sum(axis=1)
#                x1=list(range(0,100))
#                popt, pcov = curve_fit(func, x_test, y_regions)
# 
#                area1 = trapz(y_regions, dx=1)
#           
#                x3=list(range(300,10000000))
#                area3 = simps(func(x3,*popt), dx=1)
#                area=round(area1+area3)
#                area_regions[r]=area
#                    
#                    
#x_test=np.arange(0,986)
#plt.plot(x_test, S_1[x_test,:,:].sum(axis=1).sum(axis=1)) 
#plt.plot(x_test,func(x_test,*popt))             
#plt.show()
##                
#                results_sensitivity=[Number_Scenario, Name_Scenario,area,n,Copper_50, Circ]
#
#                results_sensitivity_df = pd.DataFrame(results_sensitivity)
#              
#                book=load_workbook(Project_MainPath + 'General_Results\\' +'Results_Sensitivity.xlsx')
#                
#                writer=pd.ExcelWriter(Project_MainPath + 'General_Results\\' +'Results_Sensitivity.xlsx',engine='openpyxl')
#                
#                writer.book=book
#                
#                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#                
#                
#                results_sensitivity_df.to_excel(writer, sheet_name='Sheet1',startcol=int(Number_Scenario),startrow=1,index=False)
#                
#                
#                writer.save()
#
#                EoLRR_min=min(EoLRR[:,:].sum(axis=1))      
#
#                a=np.round(S_1[:,:,:].sum(axis=1).sum(axis=1),decimals=0)
#                
#                TotalStockInTime1=[Number_Scenario, Name_Scenario]
#                
#                TotalStockInTime1.extend(a)
#                
#                TotalStockInTime_df = pd.DataFrame(TotalStockInTime1)
#                              
#                book=load_workbook(Project_MainPath + 'General_Results\\' +'Results_totalstock.xlsx')
#                
#                writer=pd.ExcelWriter(Project_MainPath + 'General_Results\\' + 'Results_totalstock.xlsx' ,engine='openpyxl')
#                
#                writer.book=book
#                
#                writer.sheets = dict((ws.title, ws) for ws in book.worksheets)
#                
#                
#                TotalStockInTime_df.to_excel(writer, sheet_name='Sheet1',startcol=int(Number_Scenario),startrow=1,index=False)
#
#
#                writer.save()


   
    
#    
#                Figurecounter = 3
#                
#                
#               # BAU_stocks[:]= S_1[:,:,:].sum(axis=1).sum(axis=1)
#                plt.figure(3)
#                
#                x = list(range(0,86))
#                Par_Time=list(range(2015,2101))
#            
#         #       plt.plot(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)/BAU_stocks[x],'k',label='BAU')
#                plt.plot(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)/BAU_stocks[x],'b',label='')
#           #     plt.plot(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)/BAU_stocks[x],'k:',label='\theta_inf')
#         #       plt.plot(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)/BAU_stocks[x],'darkgrey',marker='o',markerfacecolor='None',markevery=3,linestyle='', label='$\gamma$$_inf$')
#       #         plt.plot(Par_Time, S_1[x,:,:].sum(axis=1).sum(axis=1)/BAU_stocks[x],'darkgrey',marker='+',markerfacecolor='None',markevery=3,linestyle='', label='$\phi$$_{inf}$')
#
#                my_xticks = [Par_Time]
#                
#                                
#                plt.legend(frameon=True)
#                plt.legend(bbox_to_anchor=(1.04,1), loc="lower center")
#                xmin, xmax, ymin, ymax = 2015, 2100, 0.6, 1.42
#                print(xmin, xmax, ymin, ymax)
#                plt.axis([xmin, xmax, ymin, ymax])
#                plt.savefig(Path_Result +'sensStockplot.svg',acecolor='w',edgecolor='w', bbox_inches='tight')
#                plt.savefig(Path_Result +'sensStockplot.png',acecolor='w',edgecolor='w', bbox_inches='tight', dpi=500)
#                
#workbook  = Sensitivity_Workbook.book
#worksheet = Sensitivity_Workbook.sheets['Sheet1']
#
#chart=workbook.add_chart({'type':'line'})
## Configure the series of the chart from the dataframe data.
#chart.add_series({'values': 'results_sensitivity_df'})
#
## Insert the chart into the worksheet.
#worksheet.insert_chart('D2', chart)
## Create a chart object.
#chart = workbook.add_chart({'type': 'column'})
#
#worksheet.insert_chart('D2', chart)

#workbook  = Sensitivity_Workbook.book
#worksheet = Sensitivity_Workbook.sheets['Sheet1']
#
#chart=workbook.add_chart({'type':'line'})
## Configure the series of the chart from the dataframe data.
#chart.add_series({'values': 'results_sensitivity_df'})
#
## Insert the chart into the worksheet.
#worksheet.insert_chart('D2', chart)
## Create a chart object.
#chart = workbook.add_chart({'type': 'column'})
#
#worksheet.insert_chart('D2', chart)



#Senstitivit_Pandas_Sheet=Sensitivity_Workbook.parse('Sheet1')

#writer = pd.ExcelWriter(Path_Data +'Results_Sensitivity.xlsx', engine='xlsxwriter')
#Senstitivit_Pandas_Sheet.to_excel(writer, 'Sheet1')
#writer.save()
#
#SW=load_workbook(Path_Data +'Results_Sensitivity.xlsx')
#print(SW.get_sheet_names())
# 
# 
#Sensitivity_Workbook = xlsxwriter.Workbook(Path_Data +'Results_Sensitivity.xlsx')
#Sensitivitysheet  = Sensitivity_Workbook.add_worksheet('Sheet1')
#
#
#.get_worksheet_by_name('Sheet1')
#
#Sensitivitysheet. 
#
#Sensitivity_Workbook.save('Sensitivitysheet')
#
#sheet_by_name('Sensitivity')
#Sensitivitysheet2 = Sensitivity_Workbook.add_worksheet('flow_data')
#Sensitivitysheet.write('B2' ,    "Source")
#Sensitivity_Workbook.

#ormat to use to highlight cells.
#               bold = workbook.add_format({'bold': True})
#    
#                #Add here the flows you want to have in the Sankey Diagram 
#                FlowData = (
#                ['Source','Value','Color', 'Flow_Style','Target'],
#                ['PP', round(Global_Copper_Production2015),'(200,191,55)','ab','M'],
#                ['M',round(F_IV_A[SY+1:,:,0:5].sum()),'(0,191,255)','ab','BC'])
#
#
#                # Start from the first cell. Rows and columns are zero indexed.
#                row = 0
#                col = 0
#
#    # Iterate over the data and write it out row by row.
#    for Source,Value, Color,Flow_Style,Target in (FlowData):
#        Flowworksheet.write(row, col,     Source)
#        Flowworksheet.write(row, col + 1, Value)
#        Flowworksheet.write(row, col + 2, Color)


    
    
    
                             
  #  plt.show()

print(S_Env_Gamma[35,:,:].sum())  
print(S_Env_Omega[35,:,:].sum())      
print(S_Env_Phi[35,:].sum())     
print(S_Env_Sigma[35,:,:].sum())           
print(S_Env_Theta[35,:].sum())         
           

    
print('END')    
    #
    
print(S_Env_Omega[35,r,:].sum()+S_Env_Phi[35,r].sum()+S_Env_Sigma[35,r,:].sum()+S_Env_Theta[35,r].sum())
